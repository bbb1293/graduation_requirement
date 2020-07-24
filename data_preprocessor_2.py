#-*- coding:utf-8 -*-

import sys, io
import openpyxl, csv

def categorize_course(ws, substr, col, end_row = 300):
    output = set()

    for row in range(2, end_row + 1):
        tmp_index = 0
        val = str(ws[col + str(row)].value)

        if val == None:
            break;

        while val.find(substr, tmp_index) != -1:
            tmp_index = val.find(substr, tmp_index) + 1
            output.add(val[tmp_index - 1: tmp_index + 5])

    output = list(output)
    output.insert(0, ws[col + '1'].value)

    return output

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(
            sys.stdout.detach(), encoding = 'utf-8')
    sys.stderr = io.TextIOWrapper(
            sys.stderr.detach(), encoding = 'utf-8')

    workspace = openpyxl.load_workbook('./data_2.xlsx').active

    with open("./data_2.csv", 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)

        csvwriter.writerows([categorize_course(workspace, 'GS', ch) \
                for ch in "ABCDEFGHIJKL"])

        for code, ch in zip([code for code in ['PS', 'CH', 'BS', 
            'EC', 'MC', 'MA', 'EV'] for r in range(2)], ['M', 'N', 
                'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
                'Z']):
            csvwriter.writerow(categorize_course(workspace, code, ch))

        csvwriter.writerow(["research", "9102", "9103", "9104"])
        csvwriter.writerow(categorize_course(workspace, 'UC', 'AB'))
        csvwriter.writerow(categorize_course(workspace, 'UC', 'AC'))

        elemantary_elective = []
        for code in ['GS', 'MM', 'MD', 'CT', 'ET', 'IR']:
            elemantary_elective += categorize_course(workspace, code, 'AD')

        csvwriter.writerow(elemantary_elective)
