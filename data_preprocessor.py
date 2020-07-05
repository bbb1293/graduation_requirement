#-*- coding:utf-8 -*-

import sys, io
import openpyxl, csv

def categorize_course(ws, substr, col, end_row = 300):
    output = set()

    for row in range(2, end_row + 1):
        tmp_index = 0
        val = str(ws[col + str(row)].value)

        if val == None:
            break;

        while val.find(substr, tmp_index) != -1:
            tmp_index = val.find(substr, tmp_index) + 1
            output.add(val[tmp_index - 1: tmp_index + 5])

    output = list(output)
    output.insert(0, ws[col + '1'].value)

    return output

if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(
            sys.stdout.detach(), encoding = 'utf-8')
    sys.stderr = io.TextIOWrapper(
            sys.stderr.detach(), encoding = 'utf-8')

    workspace = openpyxl.load_workbook('./data.xlsx').active

    with open("./data.csv", 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)

        csvwriter.writerows([categorize_course(workspace, 'GS', ch) \
                for ch in "ABCDEFGHIJK"])

        elemantary_elective = []
        for code in ['GS', 'MM', 'MD', 'CT', 'ET', 'IR']:
            elemantary_elective += categorize_course(workspace, code, 'L')
        for code, ch in zip(['PS', 'CH', 'BS', 'EC', 'MC'],
                "MNOPQ"):
            elemantary_elective += categorize_course(workspace, code,
                ch, 5)

        csvwriter.writerow(elemantary_elective)
        
        csvwriter.writerows([categorize_course(workspace, 'PS', ch) \
                for ch in "RS"])
        csvwriter.writerows([categorize_course(workspace, 'CM', ch) +\
                categorize_course(workspace, 'CH', ch) for ch in "TU"])

        for code, ch in zip([code for code in ['BS', 
            'EC', 'MC', 'MA', 'EV'] for r in range(2)], ['V',
            'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE']):
            csvwriter.writerow(categorize_course(workspace, code, ch))

        csvwriter.writerow(["research", "9102", "9103", "9104"])
        csvwriter.writerow(categorize_course(workspace, 'UC', 'AG'))
